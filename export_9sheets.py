"""
export_9sheets.py
สร้าง Excel 9 sheets จาก summary object
ข้อมูลที่ไม่มีจาก ภพ.30 → แสดง template + แจ้งเตือน
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────
thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
BLUE='1A56DB'; LBLUE='EBF2FF'; WHITE='FFFFFF'; GRAY='F3F4F6'
DARK='111827'; GREEN='065F46'; LGREEN='ECFDF5'; RED='DC2626'
NAVY='1E3A5F'; AMBER='FFF9E6'; LBLUE2='DBEAFE'; DGRAY='6B7280'
ORANGE='EA580C'; LORANGE='FFF7ED'

def h(ws,r,c,v,bg=NAVY,fg=WHITE,bold=True,sz=11,center=True,merge_to=None):
    if merge_to: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=merge_to)
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name='AngsanaUPC',bold=bold,color=fg,size=sz)
    cell.fill=PatternFill('solid',fgColor=bg)
    cell.alignment=Alignment(horizontal='center' if center else 'left',
                              vertical='center',wrap_text=True)
    cell.border=bdr; return cell

def d(ws,r,c,v,bg=WHITE,fmt=None,bold=False,color=DARK,center=False,indent=0):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name='AngsanaUPC',bold=bold,color=color,size=11)
    cell.fill=PatternFill('solid',fgColor=bg)
    is_num=isinstance(v,(int,float)) and not isinstance(v,bool)
    cell.alignment=Alignment(
        horizontal='center' if center else ('right' if is_num else 'left'),
        vertical='center',indent=indent,wrap_text=True)
    cell.border=bdr
    if fmt: cell.number_format=fmt
    return cell

FMT='#,##0.00'; FMT0='#,##0'; PCT='0.00%'; PCTC='+0.00%;-0.00%'
MONTHS_TH=['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
           'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

def title3(ws,l1,l2,l3,cols=9):
    for i,(txt,sz,bold) in enumerate([(l1,16,True),(l2,14,True),(l3,11,False)],1):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=cols)
        c=ws.cell(row=i,column=1,value=txt)
        c.font=Font(name='AngsanaUPC',bold=bold,size=sz,color=DARK)
        c.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[i].height=sz+8
    ws.row_dimensions[4].height=6

def sec(ws,row,txt,cols=9,bg=NAVY):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=txt)
    c.font=Font(name='AngsanaUPC',bold=True,size=12,color=WHITE)
    c.fill=PatternFill('solid',fgColor=bg)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    ws.row_dimensions[row].height=20

def sig(ws,row,name,cols=9):
    for off,txt in enumerate([
        'ลงชื่อ ……………………………………………. หุ้นส่วนผู้จัดการ',
        f'( {name} )',
        'หมายเหตุประกอบงบการเงินเป็นส่วนหนึ่งของงบการเงินนี้'
    ]):
        r=row+off
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=cols)
        c=ws.cell(row=r,column=1,value=txt)
        c.font=Font(name='AngsanaUPC',size=11,italic=(off==2),
                    color=DGRAY if off==2 else DARK)
        c.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[r].height=20

def chgpct(ws,r,v68,v67,bg,c5=5,c6=6):
    if v68 is not None and v67 and v67!=0:
        chg=v68-v67; pct=chg/abs(v67)
        d(ws,r,c5,chg,bg,FMT,color=GREEN if chg>=0 else RED)
        cc=ws.cell(row=r,column=c6,value=pct)
        cc.font=Font(name='AngsanaUPC',size=11,color=GREEN if pct>=0 else RED)
        cc.fill=PatternFill('solid',fgColor=bg)
        cc.number_format=PCTC; cc.border=bdr
        cc.alignment=Alignment(horizontal='center',vertical='center')
    else:
        d(ws,r,c5,None,bg); d(ws,r,c6,None,bg)

def todo_banner(ws, row, msg, cols=9):
    """แถว Orange แจ้งว่าต้องกรอกข้อมูลเอง"""
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
    c=ws.cell(row=row,column=1,value=f"⚠️  {msg}")
    c.font=Font(name='AngsanaUPC',bold=True,size=11,color=WHITE)
    c.fill=PatternFill('solid',fgColor=ORANGE)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    ws.row_dimensions[row].height=22

def build_9sheets(summary, output_path,
                  manager_name="กรอกชื่อหุ้นส่วนผู้จัดการ",
                  prev_year_label="2567",
                  prev_year_data=None):
    """
    สร้าง Excel 9 sheets จาก summary
    prev_year_data: dict ข้อมูลปีก่อน เช่น
      revenue, other_income, total_revenue, cogs, admin,
      total_exp, ebit, tax, net_profit, retained_begin, retained_end
    ถ้าไม่ใส่ → คอลัมน์ปีก่อนจะว่าง
    """
    p = prev_year_data or {}
    wb = Workbook()
    co    = summary.company
    name  = co.name
    taxid = co.tax_id
    yr    = summary.year
    valid = [m for m in summary.monthly if m.total_sale > 0 or m.vat_net != 0]

    # ══════════════════════════════════════════════
    # SHEET 1: งบทดลอง
    # ══════════════════════════════════════════════
    ws1=wb.active; ws1.title=f"งบทดลอง {yr}"
    title3(ws1, name,
           f"งบทดลอง ณ วันที่ 31 ธันวาคม 25{yr}",
           f"เลขที่บัญชีจาก 1111-00 ถึง 9999-99", cols=7)
    for col,hd in enumerate(['เลขที่บัญชี','ชื่อบัญชี',
                              'ยอดยกมา\nเดบิท','ยอดยกมา\nเครดิต',
                              'ยอดคงเหลือ\nเดบิท','ยอดคงเหลือ\nเครดิต','หมายเหตุ'],1):
        h(ws1,5,col,hd); ws1.row_dimensions[5].height=36
    todo_banner(ws1, 6, "⚠️  งบทดลองต้องกรอกจากโปรแกรมบัญชี — ภพ.30 ไม่มีข้อมูลนี้", cols=7)

    # แถว template พร้อม data จาก summary ที่มี
    trial_auto = [
        ("4000-00","รายได้จากการขาย",None,None,None,sum(m.total_sale for m in valid),"จาก ภพ.30"),
        ("5000-00","ต้นทุนขาย",None,None,sum(m.purchase for m in valid),None,"จาก ภพ.30"),
        ("5300-00","ค่าใช้จ่ายบริหาร",None,None,summary.admin_expense,None,"กรอกเอง"),
        ("2135-00","ภาษีขาย",None,sum(m.vat_sale for m in valid),None,None,"จาก ภพ.30"),
        ("1154-00","ภาษีซื้อ",sum(m.vat_buy for m in valid),None,None,None,"จาก ภพ.30"),
        ("3200-00","กำไรสะสม",None,summary.retained_begin,None,None,"กรอกเอง"),
    ]
    for i,(code,acname,deb_in,crd_in,deb_bal,crd_bal,note) in enumerate(trial_auto):
        row=i+7; bg=LBLUE2 if 'จาก ภพ.30' in note else LORANGE
        ws1.row_dimensions[row].height=18
        d(ws1,row,1,code,bg,center=True,bold=True)
        d(ws1,row,2,acname,bg)
        for col,val in [(3,deb_in),(4,crd_in),(5,deb_bal),(6,crd_bal)]:
            c=ws1.cell(row=row,column=col,value=val)
            c.font=Font(name='AngsanaUPC',size=11,color=DARK)
            c.fill=PatternFill('solid',fgColor=bg)
            c.number_format=FMT; c.border=bdr
            c.alignment=Alignment(horizontal='right',vertical='center')
        d(ws1,row,7,note,bg,color=GREEN if 'ภพ.30' in note else ORANGE)

    ws1.column_dimensions['A'].width=14; ws1.column_dimensions['B'].width=38
    for col in ['C','D','E','F']: ws1.column_dimensions[col].width=20
    ws1.column_dimensions['G'].width=16

    # ══════════════════════════════════════════════
    # SHEET 2: งบดุล
    # ══════════════════════════════════════════════
    ws2=wb.create_sheet("งบดุล")
    title3(ws2, name, "งบฐานะการเงิน",
           f"ณ วันที่ 31 ธันวาคม 25{yr}", cols=6)
    for col,hd in enumerate(['รายการ','หมายเหตุข้อ',
                              f'25{yr}\n(บาท)',f'{prev_year_label}\n(บาท)',
                              'เพิ่ม(ลด)','%'],1):
        h(ws2,5,col,hd); ws2.row_dimensions[5].height=28

    total_sale_all = sum(m.total_sale for m in valid)
    total_purch    = sum(m.purchase for m in valid)

    bs=[
        ("สินทรัพย์",None,None,None,True,NAVY),
        ("สินทรัพย์หมุนเวียน",None,None,None,True,"2D4A6F"),
        ("เงินสดและรายการเทียบเท่าเงินสด","4",None,None,False,WHITE),
        ("ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น","5",None,None,False,GRAY),
        ("สินค้าคงเหลือ","",None,None,False,WHITE),
        ("สินทรัพย์หมุนเวียนอื่น","",None,None,False,GRAY),
        ("รวมสินทรัพย์หมุนเวียน","",None,None,True,LBLUE),
        ("รวมสินทรัพย์","",None,None,True,LBLUE),
        ("หนี้สินและส่วนของผู้เป็นหุ้นส่วน",None,None,None,True,NAVY),
        ("หนี้สินหมุนเวียน",None,None,None,True,"2D4A6F"),
        ("เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น","6",None,None,False,WHITE),
        ("เจ้าหนี้กรมสรรพากร-ภาษีค้างจ่าย","7",None,None,False,GRAY),
        ("รวมหนี้สินหมุนเวียน","",None,None,True,LBLUE),
        ("รวมหนี้สิน","",None,None,True,LBLUE),
        ("ส่วนของผู้เป็นหุ้นส่วน",None,None,None,True,"2D4A6F"),
        ("ทุน — [กรอกชื่อหุ้นส่วน]","",None,None,False,WHITE),
        ("รวมทุน","8",None,None,True,LBLUE),
        ("กำไร (ขาดทุน) สะสมยังไม่ได้แบ่ง","",summary.retained_end,summary.retained_begin,False,GRAY),
        ("รวมส่วนของผู้เป็นหุ้นส่วน","",None,None,True,LGREEN),
        ("รวมหนี้สินและส่วนของผู้เป็นหุ้นส่วน","",None,None,True,LGREEN),
    ]
    todo_banner(ws2, 6, "⚠️  กรอกตัวเลขงบดุลจากโปรแกรมบัญชี — เซลล์สีฟ้า = ดึงจาก ภพ.30 อัตโนมัติ", cols=6)
    for i,row_data in enumerate(bs):
        row=i+7; ws2.row_dimensions[row].height=18
        rname,note,v68,v67,is_bold,bg=row_data
        if v68 is None and not is_bold and bg not in [NAVY,"2D4A6F"]:
            bg_use = LORANGE  # ต้องกรอกเอง
        elif v68 == summary.retained_end:
            bg_use = LBLUE2   # ดึงจาก ภพ.30
        else:
            bg_use = bg
        if v68 is None and rname.startswith("ส") or v68 is None and rname.startswith("ห"):
            sec(ws2,row,rname,cols=6,bg=bg); continue
        if v68 is None and is_bold and bg in [NAVY,"2D4A6F"]:
            sec(ws2,row,rname,cols=6,bg=bg); continue
        indent=2 if not is_bold else 1
        c=ws2.cell(row=row,column=1,value=rname)
        c.font=Font(name='AngsanaUPC',bold=is_bold,size=11,color=DARK)
        c.fill=PatternFill('solid',fgColor=bg_use)
        c.alignment=Alignment(horizontal='left',vertical='center',indent=indent)
        c.border=bdr
        d(ws2,row,2,note or "",bg_use,center=True)
        for col,val in [(3,v68),(4,v67)]:
            cc=ws2.cell(row=row,column=col,value=val)
            cc.font=Font(name='AngsanaUPC',bold=is_bold,size=11,
                         color=GREEN if is_bold and val else DARK)
            cc.fill=PatternFill('solid',fgColor=bg_use)
            cc.number_format=FMT; cc.border=bdr
            cc.alignment=Alignment(horizontal='right',vertical='center')
        chgpct(ws2,row,v68,v67,bg_use)

    ws2.column_dimensions['A'].width=45; ws2.column_dimensions['B'].width=12
    for col in ['C','D','E']: ws2.column_dimensions[col].width=20
    ws2.column_dimensions['F'].width=12
    sig(ws2, len(bs)+9, manager_name, cols=6)

    # ══════════════════════════════════════════════
    # SHEET 3: PL
    # ══════════════════════════════════════════════
    ws3=wb.create_sheet("PL")
    title3(ws3, name, "งบกำไรขาดทุน",
           f"สำหรับปีสิ้นสุดวันที่ 31 ธันวาคม 25{yr}", cols=8)
    for col,hd in enumerate(['รายการ','หมายเหตุ',
                              f'25{yr}\n(บาท)',f'{prev_year_label}\n(บาท)',
                              'เพิ่ม(ลด)',f'%',
                              f'%รายได้\n25{yr}',f'%รายได้\n{prev_year_label}'],1):
        h(ws3,5,col,hd); ws3.row_dimensions[5].height=40

    rev68 = summary.total_revenue
    rev67 = p.get('revenue', p.get('total_revenue', 0))
    pl=[
        ("รายได้",None,None,None,True,NAVY),
        ("รายได้จากการขายหรือการให้บริการ","",summary.total_revenue,p.get('revenue'),False,WHITE),
        ("รายได้อื่น","9",0.,p.get('other_income'),False,GRAY),
        ("รวมรายได้","",summary.total_revenue,p.get('total_revenue'),True,LBLUE),
        ("ค่าใช้จ่าย",None,None,None,True,NAVY),
        ("ต้นทุนขายหรือต้นทุนการให้บริการ","10",summary.total_cogs,p.get('cogs'),False,WHITE),
        ("ค่าใช้จ่ายในการบริหาร","11",summary.admin_expense,p.get('admin'),False,GRAY),
        ("รวมค่าใช้จ่าย","",summary.total_cogs+summary.admin_expense,p.get('total_exp'),True,LBLUE),
        ("กำไร(ขาดทุน)ก่อนต้นทุนทางการเงินและภาษีเงินได้","",summary.ebit,p.get('ebit'),True,AMBER),
        ("ต้นทุนทางการเงิน","",0.,p.get('finance',0.),False,GRAY),
        ("กำไร(ขาดทุน)ก่อนค่าใช้จ่ายภาษีเงินได้","",summary.ebit,p.get('ebt',p.get('ebit')),True,AMBER),
        ("ภาษีเงินได้","12",summary.tax_expense,p.get('tax'),False,GRAY),
        ("กำไร(ขาดทุน)สุทธิสำหรับปี","",summary.net_profit,p.get('net_profit'),True,LGREEN),
        ("กำไร(ขาดทุน)สุทธิยกมาต้นงวด","",summary.retained_begin,p.get('retained_begin'),False,GRAY),
        ("กำไร(ขาดทุน)สุทธิยกไป","",summary.retained_end,p.get('retained_end'),True,LGREEN),
    ]
    todo_banner(ws3, 6, "⚠️  คอลัมน์ปีก่อน (2567) ต้องกรอกเอง — คอลัมน์ปีนี้ดึงจาก ภพ.30 อัตโนมัติ", cols=8)
    for i,(rname,note,v68,v67,is_total,bg) in enumerate(pl):
        row=i+7; ws3.row_dimensions[row].height=18
        if is_total and v68 is None:
            sec(ws3,row,rname,cols=8); continue
        c=ws3.cell(row=row,column=1,value=rname)
        c.font=Font(name='AngsanaUPC',bold=is_total,size=11,color=DARK)
        c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='left',vertical='center',indent=2 if not is_total else 1)
        c.border=bdr
        d(ws3,row,2,note or "",bg,center=True)
        for col,val in [(3,v68),(4,v67)]:
            cc=ws3.cell(row=row,column=col,value=val)
            cc.font=Font(name='AngsanaUPC',bold=is_total,size=11,
                         color=RED if val and val<0 else (GREEN if is_total else DARK))
            cc.fill=PatternFill('solid',fgColor=bg)
            cc.number_format=FMT; cc.border=bdr
            cc.alignment=Alignment(horizontal='right',vertical='center')
        chgpct(ws3,row,v68,v67,bg)
        d(ws3,row,7,v68/rev68 if v68 and rev68 else None,bg,PCT,center=True)
        d(ws3,row,8,v67/rev67 if v67 and rev67 else None,bg,PCT,center=True)

    ws3.column_dimensions['A'].width=48; ws3.column_dimensions['B'].width=10
    for col in ['C','D','E']: ws3.column_dimensions[col].width=18
    for col in ['F','G','H']: ws3.column_dimensions[col].width=12
    sig(ws3, len(pl)+9, manager_name, cols=8)

    # ══════════════════════════════════════════════
    # SHEET 4: หมายเหตุฯ
    # ══════════════════════════════════════════════
    ws4=wb.create_sheet("หมายเหตุฯ")
    title3(ws4, name, "หมายเหตุประกอบงบการเงิน",
           f"สำหรับปีสิ้นสุดวันที่ 31 ธันวาคม 25{yr}", cols=5)
    for col,hd in enumerate(['ข้อ','หัวข้อ','รายละเอียด',
                              f'25{yr} (บาท)',f'{prev_year_label} (บาท)'],1):
        h(ws4,5,col,hd); ws4.row_dimensions[5].height=28
    todo_banner(ws4, 6, "⚠️  กรอกรายละเอียดหมายเหตุแต่ละข้อ — ข้อมูลสีฟ้าดึงจาก ภพ.30 อัตโนมัติ", cols=5)

    notes=[
        ("1","ข้อมูลทั่วไป",f"{name}\nเลขทะเบียน: {taxid}\n[กรอกที่อยู่และวัตถุประสงค์]",None,None),
        ("2","เกณฑ์การจัดทำงบการเงิน","จัดทำตาม พ.ร.บ.วิชาชีพบัญชี พ.ศ. 2547\nใช้เกณฑ์ราคาทุนเดิม",None,None),
        ("3.1","การรับรู้รายได้และค่าใช้จ่าย","บันทึกรายได้และค่าใช้จ่ายตามเกณฑ์คงค้าง",None,None),
        ("3.2","เงินสดและรายการเทียบเท่าเงินสด","ประกอบด้วยเงินสดในมือและเงินฝากธนาคาร",None,None),
        ("3.3","สินค้าคงเหลือ","แสดงในราคาทุนหรือมูลค่าสุทธิที่จะได้รับ แล้วแต่ต่ำกว่า",None,None),
        ("4","เงินสดและรายการเทียบเท่าเงินสด","[กรอกรายละเอียดเงินสดและธนาคาร]",None,None),
        ("5","ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น","[กรอกรายละเอียด]",None,None),
        ("6","เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น","[กรอกรายละเอียด]",None,None),
        ("7","เจ้าหนี้กรมสรรพากร",f"ภาษี VAT สุทธิ: {summary.total_vat_net:,.2f} บาท",
         summary.total_vat_net,None),
        ("8","ทุน","[กรอกรายชื่อหุ้นส่วนและสัดส่วนทุน]",None,None),
        ("9","รายได้อื่น","[กรอกรายละเอียด]",0.,None),
        ("10","ต้นทุนขายและบริการ",
         f"ยอดซื้อรวม: {summary.total_cogs:,.2f} บาท\n[กรอกรายละเอียดสินค้าต้นงวด/ปลายงวด]",
         summary.total_cogs,None),
        ("11","ค่าใช้จ่ายในการบริหาร",
         f"ค่าใช้จ่ายบริหาร: {summary.admin_expense:,.2f} บาท\n[กรอกรายละเอียด]",
         summary.admin_expense,None),
        ("12","ภาษีเงินได้นิติบุคคล",
         f"กำไรก่อนภาษี: {summary.ebit:,.2f} บาท\nภาษีที่คำนวณได้: {summary.tax_expense:,.2f} บาท",
         summary.tax_expense,None),
    ]
    for i,(num,topic,detail,v68,v67) in enumerate(notes):
        row=i+7
        has_auto = v68 is not None
        bg = LBLUE2 if has_auto else (WHITE if i%2==0 else GRAY)
        ws4.row_dimensions[row].height=55
        d(ws4,row,1,num,bg,center=True,bold=True)
        d(ws4,row,2,topic,bg,bold=True)
        c=ws4.cell(row=row,column=3,value=detail)
        c.font=Font(name='AngsanaUPC',size=11)
        c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
        c.border=bdr
        for col,val in [(4,v68),(5,v67)]:
            cc=ws4.cell(row=row,column=col,value=val)
            cc.font=Font(name='AngsanaUPC',size=11,bold=(val is not None))
            cc.fill=PatternFill('solid',fgColor=bg)
            cc.number_format=FMT; cc.border=bdr
            cc.alignment=Alignment(horizontal='right',vertical='center')

    ws4.column_dimensions['A'].width=6; ws4.column_dimensions['B'].width=30
    ws4.column_dimensions['C'].width=55; ws4.column_dimensions['D'].width=18
    ws4.column_dimensions['E'].width=18

    # ══════════════════════════════════════════════
    # SHEET 5: ภพ.30 (ดึงจาก summary อัตโนมัติ)
    # ══════════════════════════════════════════════
    ws5=wb.create_sheet(f"ภพ.30.{yr}")
    title3(ws5, name, f"สรุปยอดยื่น ภ.พ.30 ปี 25{yr}",
           f"เลขประจำตัวผู้เสียภาษี: {taxid}", cols=10)
    for col,hd in enumerate(['เดือน','ยอดขาย\n(บาท)','ยอดซื้อ\n(บาท)',
                              'ภาษีขาย\n(บาท)','ภาษีซื้อ\n(บาท)',
                              'ส่วนต่าง\n(บาท)','ผลต่าง\nสะสม','สถานะ',
                              'หมายเหตุ รายได้','หมายเหตุ ต้นทุน'],1):
        h(ws5,5,col,hd); ws5.row_dimensions[5].height=40

    for i,m in enumerate(summary.monthly):
        row=i+6; bg=WHITE if i%2==0 else GRAY
        ws5.row_dimensions[row].height=18
        d(ws5,row,1,m.month_th,bg,bold=True,center=True)
        for col,val in [(2,m.sale_vat),(3,m.purchase),
                        (4,m.vat_sale),(5,m.vat_buy)]:
            c=ws5.cell(row=row,column=col,value=val)
            c.font=Font(name='AngsanaUPC',size=11); c.fill=PatternFill('solid',fgColor=bg)
            c.number_format=FMT; c.border=bdr
            c.alignment=Alignment(horizontal='right',vertical='center')
        diff=m.vat_net
        c6=ws5.cell(row=row,column=6,value=diff)
        c6.font=Font(name='AngsanaUPC',size=11,color=GREEN if diff>=0 else RED)
        c6.fill=PatternFill('solid',fgColor=bg)
        c6.number_format=FMT; c6.border=bdr
        c6.alignment=Alignment(horizontal='right',vertical='center')
        d(ws5,row,7,m.vat_cumul,bg,FMT,color=GREEN if m.vat_cumul>=0 else RED)
        status='ชำระ' if diff>=0 else 'ยกไป'
        d(ws5,row,8,status,LGREEN if diff>=0 else LBLUE,center=True,
          color=GREEN if diff>=0 else BLUE)
        d(ws5,row,9,m.sale_vat,bg,FMT0)
        d(ws5,row,10,m.purchase,bg,FMT0)

    tr5=18; ws5.row_dimensions[tr5].height=22
    h(ws5,tr5,1,"รวมทั้งปี",DARK)
    for col,val in [
        (2,sum(m.sale_vat for m in summary.monthly)),
        (3,sum(m.purchase for m in summary.monthly)),
        (4,sum(m.vat_sale for m in summary.monthly)),
        (5,sum(m.vat_buy for m in summary.monthly)),
        (6,summary.total_vat_net)
    ]:
        c=ws5.cell(row=tr5,column=col,value=val)
        c.font=Font(name='AngsanaUPC',bold=True,size=11,color=WHITE)
        c.fill=PatternFill('solid',fgColor=DARK)
        c.number_format=FMT; c.border=bdr
        c.alignment=Alignment(horizontal='right',vertical='center')
    for col in [7,8,9,10]: d(ws5,tr5,col,"",DARK)

    ws5.column_dimensions['A'].width=8
    for col in ['B','C','D','E','F','G']: ws5.column_dimensions[col].width=16
    ws5.column_dimensions['H'].width=10
    ws5.column_dimensions['I'].width=18; ws5.column_dimensions['J'].width=18

    # ══════════════════════════════════════════════
    # SHEET 6: ภาษีที่ต้องชำระ (คำนวณจาก ภพ.30)
    # ══════════════════════════════════════════════
    ws6=wb.create_sheet("ภาษีที่ต้องชำระ")
    title3(ws6, name, f"ภาษีที่ต้องชำระปี 25{yr}",
           f"เลขประจำตัวผู้เสียภาษี: {taxid}", cols=7)
    for col,hd in enumerate(['ประเภทภาษี','งวด/เดือน','ฐานภาษี (บาท)',
                              'อัตราภาษี','ภาษีที่คำนวณ (บาท)',
                              'หัก ณ ที่จ่าย (บาท)','ภาษีที่ต้องชำระ (บาท)'],1):
        h(ws6,5,col,hd); ws6.row_dimensions[5].height=36

    tax_items=[
        ("ภาษีมูลค่าเพิ่ม (ภพ.30)","รายเดือน",
         sum(m.total_sale for m in valid),0.07,
         sum(m.vat_sale for m in valid),None,summary.total_vat_net,WHITE),
        ("ภาษีหัก ณ ที่จ่าย (ภงด.1)","รายเดือน",None,0.05,None,None,None,GRAY),
        ("ภาษีหัก ณ ที่จ่าย (ภงด.3)","รายเดือน",None,0.03,None,None,None,WHITE),
        ("ภาษีเงินได้นิติบุคคลครึ่งปี (ภงด.51)",f"ส.ค. 25{yr}",None,0.20,None,None,0.,GRAY),
        ("ภาษีเงินได้นิติบุคคลประจำปี (ภงด.50)",f"พ.ค. 25{int(yr)+1}",
         summary.ebit,0.20,summary.tax_expense,None,summary.tax_expense,WHITE),
    ]
    todo_banner(ws6, 6, "⚠️  กรอก ภงด.1 / ภงด.3 และ หัก ณ ที่จ่าย เอง — แถวสีฟ้าดึงจาก ภพ.30 อัตโนมัติ", cols=7)
    for i,(rname,period,base,rate,calc,wht,net,bg) in enumerate(tax_items):
        row=i+7; ws6.row_dimensions[row].height=20
        bg_use = LBLUE2 if rname.startswith('ภาษีมูลค่าเพิ่ม') or rname.startswith('ภาษีเงินได้นิติบุคคลป') else bg
        d(ws6,row,1,rname,bg_use,bold=True)
        d(ws6,row,2,period,bg_use,center=True)
        d(ws6,row,3,base,bg_use,FMT)
        cc=ws6.cell(row=row,column=4,value=rate)
        cc.font=Font(name='AngsanaUPC',size=11); cc.fill=PatternFill('solid',fgColor=bg_use)
        cc.number_format='0.00%'; cc.border=bdr
        cc.alignment=Alignment(horizontal='center',vertical='center')
        d(ws6,row,5,calc,bg_use,FMT)
        d(ws6,row,6,wht,bg_use,FMT)
        cc=ws6.cell(row=row,column=7,value=net)
        cc.font=Font(name='AngsanaUPC',bold=True,size=11,
                     color=GREEN if (net and net>=0) else (RED if net and net<0 else ORANGE))
        cc.fill=PatternFill('solid',fgColor=LGREEN if (net and net>=0) else
                            ('FEF2F2' if (net and net<0) else LORANGE))
        cc.number_format=FMT; cc.border=bdr
        cc.alignment=Alignment(horizontal='right',vertical='center')

    tr6=len(tax_items)+7; ws6.row_dimensions[tr6].height=24
    h(ws6,tr6,1,"รวมภาษีทั้งหมดที่ต้องชำระ",DARK,merge_to=6)
    auto_total=summary.total_vat_net+summary.tax_expense
    cc=ws6.cell(row=tr6,column=7,value=auto_total)
    cc.font=Font(name='AngsanaUPC',bold=True,size=12,color=WHITE)
    cc.fill=PatternFill('solid',fgColor=BLUE); cc.number_format=FMT; cc.border=bdr
    cc.alignment=Alignment(horizontal='right',vertical='center')
    ws6.column_dimensions['A'].width=40; ws6.column_dimensions['B'].width=18
    for col in ['C','D','E','F','G']: ws6.column_dimensions[col].width=20

    # ══════════════════════════════════════════════
    # SHEET 7: เงินเดือน (template)
    # ══════════════════════════════════════════════
    ws7=wb.create_sheet("เงินเดือน")
    title3(ws7, name, f"เงินเดือนพนักงาน ปี 25{yr}",
           "ข้อมูลตามเอกสารลูกค้า", cols=16)
    headers7=['ลำดับ','ชื่อ-สกุล']+MONTHS_TH+['รวม']
    for col,hd in enumerate(headers7,1):
        h(ws7,5,col,hd); ws7.row_dimensions[5].height=28
    todo_banner(ws7, 6, "⚠️  กรอกรายชื่อพนักงานและเงินเดือนแต่ละเดือน — ภพ.30 ไม่มีข้อมูลนี้", cols=16)

    # template 5 แถว
    for i in range(1,6):
        row=i+6; bg=WHITE if i%2==0 else LORANGE
        ws7.row_dimensions[row].height=18
        d(ws7,row,1,i,bg,center=True,bold=True)
        d(ws7,row,2,f"[กรอกชื่อพนักงาน {i}]",bg,color=ORANGE)
        for j in range(3,15):
            d(ws7,row,j,None,bg)
        d(ws7,row,15,None,bg)

    ws7.column_dimensions['A'].width=8; ws7.column_dimensions['B'].width=28
    for col in range(3,16): ws7.column_dimensions[get_column_letter(col)].width=12

    # ══════════════════════════════════════════════
    # SHEET 8: สินค้าคงเหลือ (template)
    # ══════════════════════════════════════════════
    ws8=wb.create_sheet("สินค้าคงเหลือ")
    title3(ws8, name,
           f"สต็อกสินค้าคงเหลือ ณ วันที่ 31 ธันวาคม 25{yr}",
           "รายการสินค้าคงเหลือ", cols=8)
    for col,hd in enumerate(['เดือนที่ซื้อ','วันที่ซื้อ','หมายเลขเครื่อง',
                              'ทะเบียนรถ/รหัสสินค้า','ราคาทุน (บาท)',
                              'ค่าดำเนินการ (บาท)','รวมต้นทุน (บาท)','หมายเหตุ'],1):
        h(ws8,5,col,hd); ws8.row_dimensions[5].height=28
    todo_banner(ws8, 6, "⚠️  กรอกรายการสินค้าคงเหลือจากเอกสารลูกค้า — ภพ.30 ไม่มีข้อมูลนี้", cols=8)

    for i in range(1,6):
        row=i+6; bg=LORANGE if i%2==0 else WHITE
        ws8.row_dimensions[row].height=18
        d(ws8,row,1,f"[เดือน]",bg,center=True,color=ORANGE)
        d(ws8,row,2,f"[วันที่]",bg,center=True,color=ORANGE)
        d(ws8,row,3,f"[เลขเครื่อง/รหัส]",bg,color=ORANGE)
        d(ws8,row,4,f"[ทะเบียน/ชื่อสินค้า]",bg,color=ORANGE)
        d(ws8,row,5,None,bg,FMT0); d(ws8,row,6,None,bg,FMT0)
        d(ws8,row,7,None,bg,FMT0,bold=True,color=BLUE)
        d(ws8,row,8,"",bg)

    # แถวรวมสต็อก
    tr8=13; ws8.row_dimensions[tr8].height=22
    h(ws8,tr8,1,"[รวม X รายการ]",DARK,merge_to=4)
    for col in [5,6,7]: d(ws8,tr8,col,None,DARK)
    d(ws8,tr8,8,"",DARK)

    ws8.column_dimensions['A'].width=10; ws8.column_dimensions['B'].width=14
    ws8.column_dimensions['C'].width=18; ws8.column_dimensions['D'].width=28
    for col in ['E','F','G']: ws8.column_dimensions[col].width=18
    ws8.column_dimensions['H'].width=16

    # ══════════════════════════════════════════════
    # SHEET 9: ต้นทุนข้ามรอบ (template)
    # ══════════════════════════════════════════════
    ws9=wb.create_sheet("ต้นทุนข้ามรอบ")
    title3(ws9, name,
           f"รายงานต้นทุนข้ามรอบ ปี 25{yr} (ภาษีซื้อ)",
           f"สินค้าที่ซื้อในปี 25{int(yr)-1} และนำมาบันทึกต้นทุนในปี 25{yr}", cols=9)
    for col,hd in enumerate(['ที่','วันที่','หมายเลขเครื่อง','ทะเบียนรถ',
                              'เลขที่เอกสาร','ชื่อผู้ประกอบการ',
                              'รวม VAT (บาท)','VAT 7% (บาท)','มูลค่าสินค้า (บาท)'],1):
        h(ws9,5,col,hd); ws9.row_dimensions[5].height=36
    todo_banner(ws9, 6, "⚠️  กรอกรายการต้นทุนข้ามรอบจากเอกสารลูกค้า — ภพ.30 ไม่มีข้อมูลนี้", cols=9)

    for i in range(1,5):
        row=i+6; bg=LORANGE if i%2==0 else WHITE
        ws9.row_dimensions[row].height=18
        d(ws9,row,1,i,bg,center=True,bold=True)
        for col in [2,3,4,5,6]:
            d(ws9,row,col,f"[กรอก]",bg,color=ORANGE)
        for col in [7,8,9]:
            d(ws9,row,col,None,bg,FMT0)

    tr9=12; ws9.row_dimensions[tr9].height=22
    h(ws9,tr9,1,"รวม",DARK,merge_to=6)
    for col in [7,8,9]: d(ws9,tr9,col,None,DARK)

    ws9.column_dimensions['A'].width=6; ws9.column_dimensions['B'].width=14
    ws9.column_dimensions['C'].width=18; ws9.column_dimensions['D'].width=26
    ws9.column_dimensions['E'].width=16; ws9.column_dimensions['F'].width=38
    for col in ['G','H','I']: ws9.column_dimensions[col].width=18

    # ══════════════════════════════════════════════
    # Save
    # ══════════════════════════════════════════════
    wb.save(output_path)
    print(f"✅ สร้าง Excel 9 sheets: {output_path}")
    print(f"   1. งบทดลอง {yr}      — auto + template")
    print(f"   2. งบดุล              — template (กำไรสะสมอัตโนมัติ)")
    print(f"   3. PL                 — auto จาก ภพ.30")
    print(f"   4. หมายเหตุฯ          — auto + template")
    print(f"   5. ภพ.30.{yr}         — auto จาก ภพ.30 ✅")
    print(f"   6. ภาษีที่ต้องชำระ    — auto + template")
    print(f"   7. เงินเดือน          — template ⚠️  กรอกเอง")
    print(f"   8. สินค้าคงเหลือ      — template ⚠️  กรอกเอง")
    print(f"   9. ต้นทุนข้ามรอบ      — template ⚠️  กรอกเอง")
    return output_path
